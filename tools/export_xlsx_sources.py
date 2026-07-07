from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "size-chart-view.yaml"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, float):
        return f"{value:.3f}".rstrip("0").rstrip(".")
    return str(value).strip()


def parse_yaml_config(path: Path) -> dict[str, Any]:
    config: dict[str, Any] = {}
    stack: list[tuple[int, Any]] = [(-1, config)]

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        if not raw_line.strip() or raw_line.lstrip().startswith("#"):
            continue

        indent = len(raw_line) - len(raw_line.lstrip(" "))
        line = raw_line.strip()
        while stack and indent <= stack[-1][0]:
            stack.pop()

        parent = stack[-1][1]
        if line.startswith("- "):
            item_text = line[2:].strip()
            item: dict[str, Any] = {}
            parent.append(item)
            stack.append((indent, item))
            if item_text:
                key, value = split_key_value(item_text)
                item[key] = parse_scalar(value)
            continue

        key, value = split_key_value(line)
        if value == "":
            next_container: Any = [] if key.endswith("sources") else {}
            parent[key] = next_container
            stack.append((indent, next_container))
        else:
            parent[key] = parse_scalar(value)

    return config


def split_key_value(text: str) -> tuple[str, str]:
    if ":" not in text:
        return text, ""
    key, value = text.split(":", 1)
    return key.strip(), value.strip()


def parse_scalar(value: str) -> Any:
    value = value.strip().strip('"').strip("'")
    if value.lower() == "true":
        return True
    if value.lower() == "false":
        return False
    return value


def worksheet_rows(workbook: Any, sheet_name: str, header_row: int) -> tuple[list[str], list[dict[str, str]]]:
    worksheet = workbook[sheet_name]
    raw_headers = [clean(cell) for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    headers: list[str] = []
    column_headers: list[str] = []
    seen: dict[str, int] = {}
    for index, header in enumerate(raw_headers, start=1):
        if not header:
            column_headers.append("")
            continue
        if header in seen:
            seen[header] += 1
            header = f"{header}_{seen[header]}"
        else:
            seen[header] = 1
        headers.append(header)
        column_headers.append(header)

    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        row: dict[str, str] = {}
        has_value = False
        for header_index, value in enumerate(values):
            if header_index >= len(column_headers):
                break
            header = column_headers[header_index]
            if header:
                text = clean(value)
                row[header] = text
                has_value = has_value or bool(text)
        if has_value:
            rows.append(row)
    return headers, rows


def normalize_size_reference(headers: list[str], rows: list[dict[str, str]]) -> dict[str, Any]:
    normalized_rows = []
    for row in rows:
        if not clean(row.get("内部尺码")):
            continue
        item = dict(row)
        item["型号"] = clean(row.get("内部尺码"))
        normalized_rows.append(item)

    preferred = ["型号", "分类", "CAB", "长_in", "宽_in", "高_in", "通用尺码", "原长宽高_in", "长宽高_mm", "备注"]
    output_headers = [header for header in preferred if any(clean(row.get(header)) for row in normalized_rows)]
    return {"headers": output_headers, "rows": normalized_rows}


def normalize_match_rows(source: str, rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    normalized = []
    for row in rows:
        size = clean(row.get("确认尺码"))
        if not clean(row.get("MAKE")) and not clean(row.get("MODEL")):
            continue
        values = dict(row)
        values["CONST"] = clean(row.get("结构")) or clean(row.get("分类"))
        values["TYPE"] = clean(row.get("结构")) or clean(row.get("分类"))
        values["SIZE"] = size
        values["SOURCE"] = source
        normalized.append(
            {
                "make": clean(row.get("MAKE")),
                "model": clean(row.get("MODEL")),
                "year": clean(row.get("YEAR")),
                "years": expand_years(row.get("YEAR")),
                "construct": values["CONST"],
                "cab": clean(row.get("CAB")),
                "bed": clean(row.get("BED")),
                "type": values["TYPE"],
                "size": size,
                "values": values,
                "title": " ".join(part for part in [clean(row.get("MAKE")), clean(row.get("MODEL"))] if part),
                "description": "",
                "searchText": " ".join(clean(value) for value in row.values()),
                "directory": source,
                "source": source,
                "sourceTags": [source.lower()],
                "file": "",
            }
        )
    return normalized


def expand_years(value: Any) -> list[int]:
    text = clean(value)
    if not text:
        return []
    parts = text.replace("–", "-").split("-")
    if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
        start, end = int(parts[0]), int(parts[1])
        if start <= end and end - start <= 150:
            return list(range(start, end + 1))
    return [int(part) for part in text.replace("/", " ").split() if part.isdigit()]


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export frontend query JSON from an Excel workbook.")
    parser.add_argument("--xlsx-source", type=Path, help="Workbook to export. Defaults to excel_source.path in the YAML config.")
    args = parser.parse_args()

    config = parse_yaml_config(CONFIG_PATH)
    source_config = config["excel_source"]
    workbook_path = args.xlsx_source.resolve() if args.xlsx_source else ROOT / source_config["path"]
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    reference_config = config["size_reference"]
    ref_headers, ref_rows = worksheet_rows(
        workbook,
        reference_config["sheet"],
        int(reference_config.get("header_row", "2")),
    )
    write_json(ROOT / reference_config["data_path"], normalize_size_reference(ref_headers, ref_rows))

    match_payload = {"columns": [], "sources": []}
    all_columns: list[str] = []
    for source in config["match_sources"]:
        headers, rows = worksheet_rows(workbook, source["sheet"], int(source.get("header_row", "1")))
        records = normalize_match_rows(source["name"], rows)
        columns = [column for column in source.get("columns", "").split(",") if column]
        if not columns:
            columns = ["MODEL", "YEAR", "TYPE", "CAB", "BED", "SIZE"]
        for column in columns:
            if column not in all_columns:
                all_columns.append(column)
        match_payload["sources"].append(
            {
                "name": source["name"],
                "label": source.get("label", source["name"]),
                "sheet": source["sheet"],
                "columns": columns,
                "records": records,
            }
        )
    match_payload["columns"] = all_columns
    write_json(ROOT / source_config["match_data_path"], match_payload)


if __name__ == "__main__":
    main()
